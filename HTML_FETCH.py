from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import os
import shutil
from openpyxl import load_workbook
import requests
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Log-in Info
print("입력 후 엔터 눌러주세요!\n")
username = input("아이디: ")
password = input("비밀번호: ")
student_name_input = input("학생 이름: ")

# Loading takes some time
print("로딩중...")

# Find chromedriver
driver_directory = os.path.dirname(os.path.abspath(__file__))
webdriver_path = "r'" + driver_directory + "'"

# Replace these with the URL of the login page and your login credentials
login_url = 'http://erp.livedu.co/manage/bbs/login.php?url=%2Fmanage%2Findex.php%3F'

# Create a new instance of the Chrome web driver
driver = webdriver.Chrome()

# Navigate to the login page
driver.get(login_url)

# Find the username and password input fields and fill them in
username_input = driver.find_element(By.ID, 'user_id')  # Replace with the actual ID of the username input field
password_input = driver.find_element(By.ID, 'password')  # Replace with the actual ID of the password input field
checkbox = driver.find_element(By.ID, 'login_section_T')
username_input.send_keys(username)
password_input.send_keys(password)
checkbox.click()

# Find and click the login button
login_button = driver.find_element(By.CSS_SELECTOR, '#body > form > div > div.login > input')  # Replace with the
# actual ID of the login button
login_button.click()

# Wait for the page to load (you might need to adjust this wait time)
driver.implicitly_wait(10)

# Get the cookies and headers from the current page
cookies_driver = driver.get_cookies()
cookies = {cookie['name']: cookie['value'] for cookie in cookies_driver}

# Get student info via chromedriver
# Initial request for student info
student_url = 'http://erp.livedu.co/manage/student/student.php'
driver.get(student_url)

# Find link to left side student name button
link_element = driver.find_element(By.XPATH, f"//a[text()='{student_name_input}']")
onclick_attribute = link_element.get_attribute('onclick')
student_id = onclick_attribute.split("viewStudentInfo('")[1].split("')")[0]
link_element.click()

# Student phone number
time.sleep(1)
student_phone_1 = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, "st_hp1"))
)
first = student_phone_1.get_attribute('value')
student_phone_2 = driver.find_element(By.ID, 'st_hp2')
second = student_phone_2.get_attribute('value')
student_phone_3 = driver.find_element(By.ID, 'st_hp3')
third = student_phone_3.get_attribute('value')
student_phone = first + second + third

# Student gender
student_gender_check = driver.find_element(By.CSS_SELECTOR, 'input[type="radio"][name="st_sex"][value="F"]')
if student_gender_check.get_attribute('checked') == 'true':
    student_gender = '여'
else:
    student_gender = '남'

# Student school
xpath = f"//a[text()='{student_name_input}']"
student_school_path = xpath + "/../../../td[4]"
student_school_element = driver.find_element(By.XPATH, student_school_path)
student_school = student_school_element.text

# Student grade
grade_option = driver.find_element(By.CSS_SELECTOR, 'select[name="st_grade"] option[selected]')
grade_option_content = grade_option.text

# Student major
try:
    major_option = driver.find_element(By.CSS_SELECTOR, 'select[name="st_class"] option[selected]')
    option_elements = major_option.find_elements(By.TAG_NAME, "option")
    major_option_content = major_option.text

except Exception as e:
    major_option_content = "없음"

# Second request for class info
class_url = 'http://erp.livedu.co/manage/student/student_info.inc_attendance.php?st_no=' + student_id + '\''

response = requests.get(class_url, cookies=cookies, verify=False)

# Soup for class content
soup_class = BeautifulSoup(response.text, 'html.parser')

driver.quit()

data_class = soup_class.select('body > table > tr:nth-child(3) > td > form > table:nth-child(4) div')

# Pre-create items
student_subject = None
teacher_name = None
dates = []  # 수업 일자
class_content = []  # 수업 범위
homework = []  # 숙제 범위
hw_achievement = []  # 숙제 달성도
study = []  # 수업 소견

for index, tag in enumerate(data_class[11:], 1):
    if index == 2:
        student_subject = tag.contents[0]

    if index == 3:
        teacher_name = tag.contents[0]

    if index % 11 == 4:
        dates.append(tag.contents[0])

    if index % 11 == 0 and index > 21:
        study.append(tag.contents[10][18:])
    if index % 11 == 0 and index > 21:
        homework.append(tag.contents[8][12:])
    if index % 11 == 0 and index > 21:
        hw_achievement.append(tag.contents[6][17:])

    if index == 11:  # 첫줄만 에러 떠서 별도 지정
        inner = tag.contents[1].text
        x = inner.splitlines()[5]  # 수업 소견
        study.append(x[16:])
        y = inner.splitlines()[4]  # 숙제
        homework.append(y[10:])
        z = inner.splitlines()[3]  # 숙제 달성도
        hw_achievement.append(z[15:])

    if index % 11 == 10:
        class_content.append(tag.contents[0])

dates.reverse()
study.reverse()
class_content.reverse()
homework.reverse()
hw_achievement.reverse()


# Duplicate finder, true if duplicate exists
def has_duplicates(lst):
    return len(lst) != len(set(lst))


# If duplicate exists, class time is 2hrs
if has_duplicates(dates):
    class_time = 2
else:
    class_time = 1

# Total class time
total_time = len(dates) * 60


# Remove duplicates
def remove_duplicates(lst):
    return list(dict.fromkeys(lst))

# Find indices to remove for hw_a because it has multiple duplicates
seen = set()
indices_to_remove = []
for i, (num1, num2, num3, num4, num5) in enumerate(zip(dates, study, class_content, homework, hw_achievement)):
    if num1 in seen:
        indices_to_remove.append(i)
    else:
        seen.add(num1)

# Remove all duplicates
for index in reversed(indices_to_remove):
    del dates[index]
    del study[index]
    del class_content[index]
    del homework[index]
    del hw_achievement[index]

# 월말보고서 불러오기, 완성된 엑셀은 '완성소견서'
shutil.copy("소견서.xlsx", "완성소견서.xlsx")
file_source = "완성소견서.xlsx"

workbook = load_workbook(filename=file_source)

ws = workbook["월말소견서"]


# Inserting function
def enter_cell(a, b, component):
    ws.cell(row=a, column=b).value = component


# Total class time
enter_cell(12, 12, str(total_time))
# 학생 이름, 학년, 선생님 이름, 과목
enter_cell(6, 2, student_name_input)
enter_cell(6, 3, student_gender)
enter_cell(6, 5, student_school)
enter_cell(6, 12, grade_option_content)
enter_cell(6, 14, major_option_content)
enter_cell(6, 18, student_phone)
enter_cell(12, 2, teacher_name)
enter_cell(12, 3, student_subject)

# Insert remaining data according to count
count = 0
while count < len(dates):
    count_a = int(17 + 7 * count)
    enter_cell(count_a, 2, dates[count])
    enter_cell(count_a, 3, int(class_time) * 60)
    enter_cell(count_a, 5, class_content[count])
    enter_cell(count_a, 16, hw_achievement[count])
    enter_cell(count_a + 2, 3, homework[count])
    enter_cell(count_a + 4, 3, study[count])
    count += 1

workbook.save(filename=file_source)

print("완료되었습니다. 3...", end=''), time.sleep(1), print("2...", end=''), time.sleep(1), print("1...", end='')
